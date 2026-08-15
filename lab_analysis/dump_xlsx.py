import openpyxl, sys, json

def dump(path):
    print("="*100)
    print("WORKBOOK:", path)
    print("="*100)
    wb = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    print("SHEETS:", wb.sheetnames)
    # named ranges
    try:
        print("DEFINED NAMES:")
        for name, dn in wb.defined_names.items():
            print("   ", name, "->", dn.value)
    except Exception as e:
        print("defined names err", e)
    for ws in wb.worksheets:
        wsv = wbv[ws.title]
        print("\n" + "#"*90)
        print("SHEET:", ws.title, "dims:", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
        print("Hidden sheet?", ws.sheet_state)
        # hidden cols
        hidden_cols = [c for c,dim in ws.column_dimensions.items() if dim.hidden]
        hidden_rows = [r for r,dim in ws.row_dimensions.items() if dim.hidden]
        print("Hidden cols:", hidden_cols)
        print("Hidden rows:", hidden_rows)
        # data validations
        try:
            for dv in ws.data_validations.dataValidation:
                print("  DATA VALIDATION:", dv.sqref, "type=", dv.type, "formula1=", dv.formula1)
        except Exception as e:
            print("dv err", e)
        # merged
        print("Merged:", [str(m) for m in ws.merged_cells.ranges][:60])
        # cells
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                vv = wsv[cell.coordinate].value
                is_formula = isinstance(v, str) and v.startswith("=")
                if is_formula:
                    print(f"  {cell.coordinate}: FORMULA {v}  =>value= {vv!r}")
                else:
                    print(f"  {cell.coordinate}: {v!r}")

for p in sys.argv[1:]:
    try:
        dump(p)
    except Exception as e:
        print("ERROR dumping", p, e)
